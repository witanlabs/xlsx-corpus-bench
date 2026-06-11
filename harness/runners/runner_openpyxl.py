#!/usr/bin/env python3
"""openpyxl runner.

Per file, measures:
  load      — load_workbook() and touch every sheet's dimensions
  roundtrip — save to out dir, reload the saved copy with openpyxl
  recalc    — not supported (openpyxl has no calculation engine); reported N/A

Emits one JSON line per file to --out (append mode, resumable).
"""
import argparse
import json
import os
import signal
import sys
import time
import warnings

import openpyxl

PER_FILE_TIMEOUT_S = 120


class Timeout(Exception):
    pass


def _alarm(_sig, _frm):
    raise Timeout()


def process(corpus: str, rec: dict, out_dir: str) -> dict:
    path = os.path.join(corpus, rec["path"])
    keep_vba = rec["ext"] == ".xlsm"
    result = {
        "sha256": rec["sha256"],
        "path": rec["path"],
        "lib": "openpyxl",
        "load": {"ok": False, "ms": None, "error": None},
        "roundtrip": {"ok": False, "ms": None, "error": None, "out": None},
        "recalc": {"supported": False, "reason": "no calculation engine"},
    }
    wb = None
    t0 = time.monotonic()
    try:
        signal.alarm(PER_FILE_TIMEOUT_S)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path, keep_vba=keep_vba)
            for ws in wb.worksheets:
                _ = ws.calculate_dimension()
        result["load"] = {"ok": True, "ms": int((time.monotonic() - t0) * 1000), "error": None}
    except Timeout:
        result["load"]["error"] = "timeout"
        return result
    except Exception as e:
        result["load"]["error"] = f"{type(e).__name__}: {e}"[:500]
        return result
    finally:
        signal.alarm(0)

    out_path = os.path.join(out_dir, rec["sha256"] + rec["ext"])
    t0 = time.monotonic()
    try:
        signal.alarm(PER_FILE_TIMEOUT_S)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb.save(out_path)
            wb2 = openpyxl.load_workbook(out_path, keep_vba=keep_vba)
            for ws in wb2.worksheets:
                _ = ws.calculate_dimension()
        result["roundtrip"] = {
            "ok": True,
            "ms": int((time.monotonic() - t0) * 1000),
            "error": None,
            "out": out_path,
        }
    except Timeout:
        result["roundtrip"]["error"] = "timeout"
    except Exception as e:
        result["roundtrip"]["error"] = f"{type(e).__name__}: {e}"[:500]
        result["roundtrip"]["out"] = out_path if os.path.exists(out_path) else None
    finally:
        signal.alarm(0)
    return result


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--corpus", required=True)
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--out", required=True, help="results jsonl (appended)")
    ap.add_argument("--out-dir", required=True, help="dir for round-tripped files")
    args = ap.parse_args()

    signal.signal(signal.SIGALRM, _alarm)
    os.makedirs(args.out_dir, exist_ok=True)

    done = set()
    if os.path.exists(args.out):
        with open(args.out) as f:
            for line in f:
                try:
                    done.add(json.loads(line)["sha256"])
                except Exception:
                    pass

    with open(args.manifest) as mf, open(args.out, "a") as out:
        for line in mf:
            rec = json.loads(line)
            if rec["sha256"] in done:
                continue
            res = process(args.corpus, rec, args.out_dir)
            out.write(json.dumps(res) + "\n")
            out.flush()
    print(f"openpyxl {openpyxl.__version__} done", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
